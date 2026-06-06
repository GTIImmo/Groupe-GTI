from __future__ import annotations

import json
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX_PATH = ROOT / "liste mandat du 02_02_2026 au 28_02_2026.xlsx"

EXCLUSIVITY_LABELS = {
    "1": "EXCLUSIF",
    "3": "ACCORD",
    "5": "SIMPLE",
}


def _normalize_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _normalize_date(value: Any) -> str | None:
    text = _normalize_text(value)
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        return text[:10]


def _build_montant(linked_product_ref: dict[str, Any], fees_value: Any) -> str | None:
    price = _normalize_text(linked_product_ref.get("price"))
    if price and price not in {"0", "0.0"}:
        return price

    seller_net = _normalize_text(linked_product_ref.get("price_seller_net"))
    fees = _normalize_text(fees_value)
    if seller_net and seller_net.isdigit() and fees and fees.isdigit():
        return str(int(seller_net) + int(fees))
    if seller_net:
        return seller_net
    return None


def _extract_mandate_number_from_data(data: dict[str, Any]) -> str | None:
    key_data = data.get("keyData") or {}
    if isinstance(key_data, dict):
        numero = _normalize_text(key_data.get("NO_MANDAT"))
        if numero:
            return numero

    mandat_bloc = data.get("mandat_mandatdispo") or {}
    props = mandat_bloc.get("props") if isinstance(mandat_bloc, dict) else {}
    if isinstance(props, dict):
        no_dossier = props.get("NO_DOSSIER")
        numero = _normalize_text((no_dossier or {}).get("value") if isinstance(no_dossier, dict) else None)
        if numero and numero.isdigit():
            return numero

    return None


@lru_cache(maxsize=1)
def load_manual_mandat_corrections(xlsx_path: Path = DEFAULT_XLSX_PATH) -> dict[str, dict[str, Any]]:
    if not xlsx_path.exists():
        return {}

    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [_normalize_text(value) or "" for value in rows[0]]
    index = {name: position for position, name in enumerate(header)}
    corrections: dict[str, dict[str, Any]] = {}

    for raw_row in rows[1:]:
        if not any(value is not None for value in raw_row):
            continue

        mandate_number = _normalize_text(raw_row[index["mandate_number"]])
        if not mandate_number:
            continue

        linked_product_ref_raw = _normalize_text(raw_row[index["linked_product_ref"]])
        try:
            linked_product_ref = json.loads(linked_product_ref_raw) if linked_product_ref_raw else {}
        except json.JSONDecodeError:
            linked_product_ref = {}

        exclusivity = _normalize_text(raw_row[index["exclusivity"]])
        mandate_type = EXCLUSIVITY_LABELS.get(exclusivity, "SIMPLE")
        contact_name = _normalize_text(raw_row[index["contact_full_name"]])
        contact_address = _normalize_text(raw_row[index["contact_full_address"]])
        mandants = " ".join(part for part in (contact_name, contact_address) if part)

        correction = {
            "id": f"manual:{mandate_number}",
            "numero": mandate_number,
            "type": mandate_type,
            "dateEnregistrement": _normalize_date(raw_row[index["date"]]),
            "debut": _normalize_date(raw_row[index["date_start"]]),
            "fin": _normalize_date(raw_row[index["date_end"]]),
            "cloture": None,
            "montant": _build_montant(linked_product_ref, raw_row[index["fees"]]),
            "mandants": mandants or None,
            "note": f"manual_xlsx:{xlsx_path.name}",
            "avenants": [],
            "manual_import": True,
            "linked_product_ref": _normalize_text(linked_product_ref.get("product_ref")),
            "product_recap": _normalize_text(raw_row[index["product_recap"]]),
        }
        corrections[mandate_number] = correction

    return corrections


def get_manual_mandat_correction(mandate_number: Any) -> dict[str, Any] | None:
    numero = _normalize_text(mandate_number)
    if not numero:
        return None
    correction = load_manual_mandat_corrections().get(numero)
    if correction:
        return dict(correction)
    return None


def inject_manual_mandat_if_missing(data: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(data, dict):
        return data

    mandats = data.get("mandats")
    if isinstance(mandats, list) and mandats:
        return data

    correction = get_manual_mandat_correction(mandate_number=_extract_mandate_number_from_data(data))
    if not correction:
        return data

    patched = dict(data)
    patched["mandats"] = [correction]
    return patched
